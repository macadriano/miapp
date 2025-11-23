import logging
from datetime import datetime

from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from .models import Equipo, Posicion, CatMovil, TipoEquipoGPS, ConfiguracionReceptor, EstadisticasRecepcion
from moviles.models import Movil, MovilStatus, MovilGeocode, MovilObservacion, MovilFoto, MovilNota
from .serializers import (
    EquipoSerializer,
    PosicionSerializer,
    PosicionRecorridoSerializer,
    CatMovilSerializer,
    TipoEquipoGPSSerializer,
    ConfiguracionReceptorSerializer,
    EstadisticasRecepcionSerializer,
    RecorridoStatsSerializer,
    RecorridoFiltrosSerializer,
)
from moviles.serializers import (
    MovilSerializer, MovilStatusSerializer, MovilGeocodeSerializer,
    MovilObservacionSerializer, MovilFotoSerializer, MovilNotaSerializer
)

logger = logging.getLogger(__name__)


# MovilViewSet, MovilStatusViewSet, MovilGeocodeViewSet, MovilObservacionViewSet, MovilFotoViewSet, MovilNotaViewSet
# se movieron a moviles/views.py


class EquipoViewSet(viewsets.ModelViewSet):
    queryset = Equipo.objects.all()
    serializer_class = EquipoSerializer
    permission_classes = [AllowAny]  # Permitir acceso sin autenticación temporalmente
    
    @action(detail=False, methods=['get'])
    def sin_asignar(self, request):
        """Obtener equipos disponibles (sin asignar a ningún móvil)
        
        Parámetros:
        - movil_id (opcional): ID del móvil que se está editando.
                             Si se proporciona, el equipo actual de ese móvil
                             también se incluirá en la lista.
        """
        # Obtener parámetro opcional de móvil actual
        movil_id = request.query_params.get('movil_id', None)
        gps_id_actual = None
        
        if movil_id:
            try:
                movil_actual = Movil.objects.get(id=movil_id)
                gps_id_actual = movil_actual.gps_id
            except Movil.DoesNotExist:
                pass
        
        # Buscar equipos que no estén asignados a ningún móvil
        equipos_sin_asignar = []
        
        # Obtener todos los equipos
        todos_equipos = Equipo.objects.all()
        
        for equipo in todos_equipos:
            # Verificar si este IMEI está en uso en algún móvil
            if not Movil.objects.filter(gps_id=equipo.imei).exists():
                equipos_sin_asignar.append(equipo)
            # Si estamos editando un móvil y este es su equipo actual, incluirlo
            elif movil_id and gps_id_actual and equipo.imei == gps_id_actual:
                equipos_sin_asignar.append(equipo)
        
        serializer = self.get_serializer(equipos_sin_asignar, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def asignar_movil(self, request, pk=None):
        """Asignar este equipo a un móvil específico"""
        equipo = self.get_object()
        movil_id = request.data.get('movil_id')
        
        if not movil_id:
            return Response({'error': 'movil_id es requerido'}, status=400)
        
        try:
            movil = Movil.objects.get(id=movil_id)
        except Movil.DoesNotExist:
            return Response({'error': 'Móvil no encontrado'}, status=404)
        
        # Verificar si el equipo ya está asignado
        if Movil.objects.filter(gps_id=equipo.imei).exists():
            return Response({'error': 'Este equipo ya está asignado a otro móvil'}, status=400)
        
        # Asignar el equipo al móvil
        movil.gps_id = equipo.imei
        movil.save()
        
        return Response({
            'mensaje': f'Equipo {equipo.imei} asignado correctamente al móvil {movil.patente or movil.alias}',
            'equipo': self.get_serializer(equipo).data,
            'movil': MovilSerializer(movil).data
        })
    
    @action(detail=True, methods=['post'])
    def desasignar(self, request, pk=None):
        """Desasignar este equipo del móvil actual"""
        equipo = self.get_object()
        
        try:
            movil = Movil.objects.get(gps_id=equipo.imei)
            movil.gps_id = None
            movil.save()
            
            return Response({
                'mensaje': f'Equipo {equipo.imei} desasignado correctamente del móvil {movil.patente or movil.alias}',
                'equipo': self.get_serializer(equipo).data
            })
        except Movil.DoesNotExist:
            return Response({'error': 'Este equipo no está asignado a ningún móvil'}, status=400)


# Vista para servir el frontend de móviles
def moviles_frontend(request):
    return render(request, 'moviles/index.html', {
        'default_section': 'moviles'
    })


# Vista para servir el frontend de equipos
def equipos_frontend(request):
    return render(request, 'equipos/index.html')


# Nuevos ViewSets para la estructura optimizada
# class MovilStatusViewSet(viewsets.ModelViewSet):  # MOVIDO A moviles/views.py
#     queryset = MovilStatus.objects.all()
#     serializer_class = MovilStatusSerializer


# class MovilGeocodeViewSet(viewsets.ModelViewSet):  # MOVIDO A moviles/views.py
#     queryset = MovilGeocode.objects.all()
#     serializer_class = MovilGeocodeSerializer


class PosicionViewSet(viewsets.ModelViewSet):
    queryset = Posicion.objects.none()
    serializer_class = PosicionSerializer
    
    def get_queryset(self):
        """Optimizar consultas de posiciones"""
        return Posicion.objects.filter(is_valid=True).select_related('movil').only(
            'id', 'movil', 'fec_gps', 'lat', 'lon', 'velocidad', 'rumbo', 'sats', 'hdop', 'ign_on', 'direccion'
        )
    
    @action(detail=False, methods=['get'])
    def por_movil(self, request):
        """Obtener posiciones de un móvil específico"""
        movil_id = request.query_params.get('movil_id')
        if movil_id:
            queryset = Posicion.objects.filter(
                movil_id=movil_id, 
                is_valid=True
            ).select_related('movil').only(
                'id', 'movil', 'fec_gps', 'lat', 'lon', 'velocidad', 'rumbo', 'sats', 'hdop', 'ign_on', 'direccion'
            )[:1000]  # Limitar resultados
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        return Response([])


class CatMovilViewSet(viewsets.ModelViewSet):
    queryset = CatMovil.objects.all()
    serializer_class = CatMovilSerializer


class TipoEquipoGPSViewSet(viewsets.ModelViewSet):
    queryset = TipoEquipoGPS.objects.all()
    serializer_class = TipoEquipoGPSSerializer


class ConfiguracionReceptorViewSet(viewsets.ModelViewSet):
    queryset = ConfiguracionReceptor.objects.all()
    serializer_class = ConfiguracionReceptorSerializer
    permission_classes = [AllowAny]
    
    def list(self, request, *args, **kwargs):
        """Asegurar que siempre devuelva una lista"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class EstadisticasRecepcionViewSet(viewsets.ModelViewSet):
    queryset = EstadisticasRecepcion.objects.all()
    serializer_class = EstadisticasRecepcionSerializer
    permission_classes = [AllowAny]
    
    def list(self, request, *args, **kwargs):
        """Obtener estadísticas. Si no hay datos en EstadisticasRecepcion, generar desde las posiciones."""
        queryset = self.get_queryset()
        
        # Filtrar por receptor si se proporciona el parámetro receptor
        receptor_id = request.query_params.get('receptor', None)
        if receptor_id is not None:
            try:
                receptor = ConfiguracionReceptor.objects.get(id=receptor_id)
                queryset = queryset.filter(receptor=receptor)
            except ConfiguracionReceptor.DoesNotExist:
                pass  # Si no existe, devolver vacío
        
        # Si no hay estadísticas registradas, generarlas desde las posiciones
        # PERO solo si NO se está filtrando por receptor específico
        if not queryset.exists() and receptor_id is None:
            from django.utils import timezone
            from datetime import timedelta
            from django.db.models import Count, Q
            
            # Generar estadísticas de los últimos 7 días
            fecha_inicio = timezone.now().date() - timedelta(days=7)
            
            # Obtener estadísticas por día
            posiciones = Posicion.objects.filter(
                fec_gps__date__gte=fecha_inicio
            ).order_by('fec_gps')
            
            # Agrupar por fecha
            stats_por_dia = {}
            
            for posicion in posiciones:
                fecha_key = posicion.fec_gps.date().isoformat()
                
                if fecha_key not in stats_por_dia:
                    stats_por_dia[fecha_key] = {
                        'receptor_nombre': 'Receptor TCP',
                        'fecha': fecha_key,
                        'equipos_conectados': set(),
                        'datos_recibidos': 0,
                        'datos_procesados': 0,
                        'errores': 0
                    }
                
                # Agregar móvil conectado
                if posicion.movil_id:
                    stats_por_dia[fecha_key]['equipos_conectados'].add(posicion.movil_id)
                
                # Contar datos
                stats_por_dia[fecha_key]['datos_recibidos'] += 1
                
                if posicion.is_valid:
                    stats_por_dia[fecha_key]['datos_procesados'] += 1
                else:
                    stats_por_dia[fecha_key]['errores'] += 1
            
            # Convertir a lista y convertir sets a counts
            resultado = []
            for fecha_key, stats in stats_por_dia.items():
                resultado.append({
                    'receptor_nombre': stats['receptor_nombre'],
                    'fecha': stats['fecha'],
                    'equipos_conectados': len(stats['equipos_conectados']),
                    'datos_recibidos': stats['datos_recibidos'],
                    'datos_procesados': stats['datos_procesados'],
                    'errores': stats['errores'],
                    'latencia_promedio': None
                })
            
            return Response(resultado)
        
        # Si hay estadísticas, retornarlas normalmente
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# Nuevos ViewSets para observaciones, fotos y notas
# class MovilObservacionViewSet(viewsets.ModelViewSet):  # MOVIDO A moviles/views.py
#     queryset = MovilObservacion.objects.all()
#     serializer_class = MovilObservacionSerializer
#     permission_classes = [AllowAny]
    
#     def get_queryset(self):  # MOVIDO A moviles/views.py
#         """Filtrar observaciones por móvil si se proporciona el parámetro movil_id"""
#         queryset = MovilObservacion.objects.all()
#         movil_id = self.request.query_params.get('movil_id', None)
#         if movil_id is not None:
#             queryset = queryset.filter(movil_id=movil_id)
#         return queryset


# class MovilFotoViewSet(viewsets.ModelViewSet):  # MOVIDO A moviles/views.py
#     queryset = MovilFoto.objects.all()
#     serializer_class = MovilFotoSerializer
#     permission_classes = [AllowAny]
    
#     def get_queryset(self):  # MOVIDO A moviles/views.py
#         """Filtrar fotos por móvil si se proporciona el parámetro movil_id"""
#         queryset = MovilFoto.objects.all()
#         movil_id = self.request.query_params.get('movil_id', None)
#         if movil_id is not None:
#             queryset = queryset.filter(movil_id=movil_id)
#         return queryset
    
#     @action(detail=False, methods=['post'])  # MOVIDO A moviles/views.py
#     def subir_multiple(self, request):
#         """Subir múltiples fotos a la vez"""
#         movil_id = request.data.get('movil_id')
#         if not movil_id:
#             return Response({'error': 'movil_id es requerido'}, status=400)
#         
#         try:
#             movil = Movil.objects.get(id=movil_id)
#         except Movil.DoesNotExist:
#             return Response({'error': 'Móvil no encontrado'}, status=404)
#         
#         # Verificar límite de fotos (máximo 20)
#         fotos_actuales = MovilFoto.objects.filter(movil=movil).count()
#         archivos = request.FILES.getlist('imagenes')
#         
#         if fotos_actuales + len(archivos) > 20:
#             return Response({
#                 'error': f'Límite de 20 fotos excedido. Actualmente: {fotos_actuales}, intentando subir: {len(archivos)}'
#             }, status=400)
#         
#         fotos_creadas = []
#         for archivo in archivos:
#             foto_data = {
#                 'movil': movil_id,
#                 'imagen': archivo,
#                 'titulo': request.data.get('titulo', ''),
#                 'descripcion': request.data.get('descripcion', ''),
#                 'categoria': request.data.get('categoria', 'general'),
#             }
#             
#             serializer = self.get_serializer(data=foto_data)
#             if serializer.is_valid():
#                 foto = serializer.save()
#                 fotos_creadas.append(serializer.data)
#             else:
#                 return Response(serializer.errors, status=400)
#         
#         return Response({
#             'mensaje': f'{len(fotos_creadas)} fotos subidas correctamente',
#             'fotos': fotos_creadas
#         }, status=201)


# class MovilNotaViewSet(viewsets.ModelViewSet):  # MOVIDO A moviles/views.py
#     queryset = MovilNota.objects.all()
#     serializer_class = MovilNotaSerializer
#     permission_classes = [AllowAny]
    
#     def get_queryset(self):  # MOVIDO A moviles/views.py
#         """Filtrar notas por móvil si se proporciona el parámetro movil_id"""
#         queryset = MovilNota.objects.all()
#         movil_id = self.request.query_params.get('movil_id', None)
#         if movil_id is not None:
#             queryset = queryset.filter(movil_id=movil_id)
#         return queryset


class RecorridosViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de recorridos GPS (optimizado para grandes volúmenes)."""

    queryset = Posicion.objects.none()
    serializer_class = PosicionRecorridoSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        """Filtrar posiciones según parámetros de consulta."""
        movil_id_param = self.request.query_params.get('movil_id')
        if not movil_id_param:
            return Posicion.objects.none()

        try:
            movil_id = int(movil_id_param)
        except (TypeError, ValueError):
            return Posicion.objects.none()

        queryset = (
            Posicion.objects.filter(is_valid=True, movil_id=movil_id)
            .select_related('movil')
            .defer(
                'raw_payload',
                'inputs_mask',
                'outputs_mask',
                'msg_uid',
                'seq',
                'provider',
                'protocol',
                'batt_mv',
                'ext_pwr_mv',
                'accuracy_m',
                'evento',
                'evt_tipo_id',
                'fec_report',
            )
            .order_by('fec_gps')
        )

        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fec_gps__gte=fecha_desde)

        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fec_gps__lte=fecha_hasta)

        velocidad_min = self.request.query_params.get('velocidad_min')
        if velocidad_min:
            queryset = queryset.filter(velocidad__gte=velocidad_min)

        velocidad_max = self.request.query_params.get('velocidad_max')
        if velocidad_max:
            queryset = queryset.filter(velocidad__lte=velocidad_max)

        solo_detenciones = self.request.query_params.get('solo_detenciones')
        if solo_detenciones == 'true':
            queryset = queryset.filter(velocidad__lte=5)

        solo_movimiento = self.request.query_params.get('solo_movimiento')
        if solo_movimiento == 'true':
            queryset = queryset.filter(velocidad__gt=5)

        ignicion_encendida = self.request.query_params.get('ignicion_encendida')
        if ignicion_encendida == 'true':
            queryset = queryset.filter(ign_on=True)

        vista_mapa = self.request.query_params.get('vista_mapa')
        limite_default = 50000 if vista_mapa == 'true' else 10000
        limite_maximo = 100000 if vista_mapa == 'true' else 50000

        if vista_mapa == 'true':
            self.pagination_class = None

        limite = self.request.query_params.get('limite', limite_default)
        try:
            limite = int(limite)
            limite = min(max(limite, 1), limite_maximo)
        except (TypeError, ValueError):
            limite = limite_default

        return queryset[:limite]

    @action(detail=False, methods=['post'])
    def filtrar_recorrido(self, request):
        """Endpoint para filtrar posiciones con parámetros avanzados"""
        serializer = RecorridoFiltrosSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        
        filtros = serializer.validated_data
        queryset = Posicion.objects.filter(is_valid=True)
        
        # Aplicar filtros
        if filtros.get('movil_id'):
            queryset = queryset.filter(movil_id=filtros['movil_id'])
        
        if filtros.get('fecha_desde'):
            queryset = queryset.filter(fec_gps__gte=filtros['fecha_desde'])
        
        if filtros.get('fecha_hasta'):
            queryset = queryset.filter(fec_gps__lte=filtros['fecha_hasta'])
        
        if filtros.get('velocidad_min'):
            queryset = queryset.filter(velocidad__gte=filtros['velocidad_min'])
        
        if filtros.get('velocidad_max'):
            queryset = queryset.filter(velocidad__lte=filtros['velocidad_max'])
        
        if filtros.get('solo_detenciones'):
            queryset = queryset.filter(velocidad__lte=5)
        
        if filtros.get('solo_movimiento'):
            queryset = queryset.filter(velocidad__gt=5)
        
        if filtros.get('ignicion_encendida') is not None:
            queryset = queryset.filter(ign_on=filtros['ignicion_encendida'])
        
        # Ordenar por fecha GPS
        queryset = queryset.order_by('fec_gps')
        
        # Limitar resultados
        limite = filtros.get('limite_registros', 10000)
        queryset = queryset[:limite]
        
        # Serializar resultados
        serializer = PosicionRecorridoSerializer(queryset, many=True)
        
        return Response({
            'posiciones': serializer.data,
            'total_registros': queryset.count(),
            'filtros_aplicados': filtros
        })
    
    @action(detail=False, methods=['get'])
    def estadisticas_recorrido(self, request):
        """Obtener estadísticas de un recorrido"""
        movil_id = request.query_params.get('movil_id')
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        
        if not movil_id or not fecha_desde or not fecha_hasta:
            return Response({
                'error': 'Se requieren parámetros: movil_id, fecha_desde, fecha_hasta'
            }, status=400)
        
        try:
            from django.utils import timezone
            from datetime import datetime, timedelta
            import math
            
            # Obtener posiciones del recorrido
            posiciones = Posicion.objects.filter(
                movil_id=movil_id,
                fec_gps__gte=fecha_desde,
                fec_gps__lte=fecha_hasta,
                is_valid=True
            ).order_by('fec_gps')
            
            if not posiciones.exists():
                return Response({
                    'error': 'No se encontraron posiciones para el recorrido especificado'
                }, status=404)
            
            # Calcular estadísticas básicas
            primera_posicion = posiciones.first()
            ultima_posicion = posiciones.last()
            
            # Duración en minutos
            duracion = (ultima_posicion.fec_gps - primera_posicion.fec_gps).total_seconds() / 60
            
            # Velocidades
            velocidades = [p.velocidad for p in posiciones if p.velocidad is not None]
            velocidad_maxima = max(velocidades) if velocidades else 0
            velocidad_promedio = sum(velocidades) / len(velocidades) if velocidades else 0
            
            # Distancia total (cálculo aproximado)
            distancia_total = 0
            posicion_anterior = None
            
            for posicion in posiciones:
                if posicion_anterior and posicion.lat and posicion.lon and posicion_anterior.lat and posicion_anterior.lon:
                    # Fórmula de Haversine para calcular distancia entre dos puntos
                    lat1, lon1 = float(posicion_anterior.lat), float(posicion_anterior.lon)
                    lat2, lon2 = float(posicion.lat), float(posicion.lon)
                    
                    R = 6371  # Radio de la Tierra en km
                    dlat = math.radians(lat2 - lat1)
                    dlon = math.radians(lon2 - lon1)
                    a = math.sin(dlat/2) * math.sin(dlat/2) + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2) * math.sin(dlon/2)
                    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
                    distancia = R * c
                    distancia_total += distancia
                
                posicion_anterior = posicion
            
            # Detenciones (velocidad <= 5 km/h)
            detenciones = posiciones.filter(velocidad__lte=5).count()
            
            # Tiempo detenido vs tiempo en movimiento
            tiempo_detenido = posiciones.filter(velocidad__lte=5).count()  # Aproximación
            tiempo_movimiento = posiciones.count() - tiempo_detenido
            
            # Rango de velocidades
            rango_velocidades = {
                'detenido': posiciones.filter(velocidad__lte=5).count(),
                'muy_lento': posiciones.filter(velocidad__gt=5, velocidad__lte=15).count(),
                'lento': posiciones.filter(velocidad__gt=15, velocidad__lte=30).count(),
                'moderado': posiciones.filter(velocidad__gt=30, velocidad__lte=50).count(),
                'rapido': posiciones.filter(velocidad__gt=50, velocidad__lte=80).count(),
                'muy_rapido': posiciones.filter(velocidad__gt=80).count(),
            }
            
            # Información del móvil
            movil = Movil.objects.get(id=movil_id) if Movil.objects.filter(id=movil_id).exists() else None
            
            estadisticas = {
                'movil_id': movil_id,
                'movil_info': {
                    'id': movil.id,
                    'patente': movil.patente,
                    'alias': movil.alias,
                    'codigo': movil.codigo,
                    'marca': movil.marca,
                    'modelo': movil.modelo
                } if movil else None,
                'fecha_inicio': primera_posicion.fec_gps,
                'fecha_fin': ultima_posicion.fec_gps,
                'duracion_minutos': round(duracion, 2),
                'distancia_km': round(distancia_total, 2),
                'velocidad_maxima': velocidad_maxima,
                'velocidad_promedio': round(velocidad_promedio, 2),
                'puntos_gps': posiciones.count(),
                'detenciones': detenciones,
                'tiempo_detenido_minutos': round(tiempo_detenido, 2),
                'tiempo_movimiento_minutos': round(tiempo_movimiento, 2),
                'eficiencia_combustible': None,  # Por implementar
                'rango_velocidades': rango_velocidades,
                'estadisticas_por_hora': []  # Por implementar
            }
            
            serializer = RecorridoStatsSerializer(estadisticas)
            return Response(serializer.data)
            
        except Exception as e:
            return Response({
                'error': f'Error calculando estadísticas: {str(e)}'
            }, status=500)
    
    @action(detail=False, methods=['get'])
    def exportar_recorrido(self, request):
        """Exportar recorrido en formato JSON para reproducción"""
        movil_id = request.query_params.get('movil_id')
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        
        if not movil_id or not fecha_desde or not fecha_hasta:
            return Response({
                'error': 'Se requieren parámetros: movil_id, fecha_desde, fecha_hasta'
            }, status=400)
        
        # Obtener posiciones del recorrido
        posiciones = Posicion.objects.filter(
            movil_id=movil_id,
            fec_gps__gte=fecha_desde,
            fec_gps__lte=fecha_hasta,
            is_valid=True
        ).order_by('fec_gps')
        
        # Serializar para reproducción
        datos_recorrido = []
        for posicion in posiciones:
            if posicion.lat and posicion.lon:
                datos_recorrido.append({
                    'timestamp': posicion.fec_gps.isoformat(),
                    'lat': float(posicion.lat),
                    'lon': float(posicion.lon),
                    'velocidad': posicion.velocidad or 0,
                    'rumbo': posicion.rumbo or 0,
                    'altitud': posicion.altitud or 0,
                    'ignicion': posicion.ign_on,
                    'satelites': posicion.sats or 0,
                    'calidad': posicion.calidad_senal
                })
        
        return Response({
            'recorrido': datos_recorrido,
            'total_puntos': len(datos_recorrido),
            'parametros': {
                'movil_id': movil_id,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta
            }
        })
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exportar recorrido a Excel"""
        movil_id = request.query_params.get('movil_id')
        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        
        if not movil_id or not fecha_desde or not fecha_hasta:
            return Response({
                'error': 'Se requieren los parámetros movil_id, fecha_desde y fecha_hasta'
            }, status=400)
        
        try:
            posiciones_queryset = Posicion.objects.filter(
                movil_id=movil_id,
                fec_gps__gte=fecha_desde,
                fec_gps__lte=fecha_hasta,
                is_valid=True
            ).order_by('fec_gps')
            
            if not posiciones_queryset.exists():
                return Response({'error': 'No hay posiciones para exportar con los filtros indicados'}, status=404)
            
            posiciones = []
            for posicion in posiciones_queryset:
                posiciones.append({
                    'timestamp': posicion.fec_gps.isoformat() if posicion.fec_gps else '',
                    'lat': float(posicion.lat) if posicion.lat else None,
                    'lon': float(posicion.lon) if posicion.lon else None,
                    'direccion': posicion.direccion or '',
                    'velocidad': posicion.velocidad or 0,
                    'rumbo': posicion.rumbo or 0,
                    'altitud': posicion.altitud or 0,
                    'satelites': posicion.sats or 0,
                    'ignicion': posicion.ign_on,
                    'calidad': posicion.calidad_senal or ''
                })
            
            estadisticas_response = self.estadisticas_recorrido(request)
            if isinstance(estadisticas_response.data, dict) and 'error' in estadisticas_response.data:
                estadisticas = {}
            else:
                estadisticas = estadisticas_response.data if hasattr(estadisticas_response, 'data') else estadisticas_response
            
            movil = Movil.objects.filter(id=movil_id).first()
            movil_info = {
                'patente': movil.patente if movil else 'N/A',
                'alias': movil.alias if movil else 'N/A',
                'marca': movil.marca if movil else 'N/A',
                'modelo': movil.modelo if movil else 'N/A'
            }
        except Exception as e:
            logger.exception("Error obteniendo datos para exportar Excel")
            return Response({'error': f'No se pudieron obtener los datos del recorrido: {str(e)}'}, status=500)
        
        if not posiciones:
            return Response({'error': 'No hay posiciones para exportar'}, status=400)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            import io
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recorrido GPS"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Encabezados
            headers = [
                'N°', 'Fecha/Hora', 'Latitud', 'Longitud', 'Dirección',
                'Velocidad (km/h)', 'Rumbo (°)', 'Altitud (m)', 'Satélites',
                'Encendido', 'Calidad Señal'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Datos
            for row, posicion in enumerate(posiciones, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=posicion.get('timestamp', ''))
                ws.cell(row=row, column=3, value=posicion.get('lat', ''))
                ws.cell(row=row, column=4, value=posicion.get('lon', ''))
                ws.cell(row=row, column=5, value=posicion.get('direccion', ''))
                ws.cell(row=row, column=6, value=posicion.get('velocidad', 0))
                ws.cell(row=row, column=7, value=posicion.get('rumbo', 0))
                ws.cell(row=row, column=8, value=posicion.get('altitud', 0))
                ws.cell(row=row, column=9, value=posicion.get('satelites', 0))
                ws.cell(row=row, column=10, value='Sí' if posicion.get('ignicion') else 'No')
                ws.cell(row=row, column=11, value=posicion.get('calidad', ''))
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Agregar hoja de estadísticas si están disponibles
            if estadisticas or movil_info:
                ws_stats = wb.create_sheet("Información del Recorrido")
                
                # Encabezado de estadísticas
                ws_stats.cell(row=1, column=1, value="Información")
                ws_stats.cell(row=1, column=2, value="Valor")
                
                # Aplicar estilos al encabezado
                header_cell = ws_stats.cell(row=1, column=1)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                
                header_cell2 = ws_stats.cell(row=1, column=2)
                header_cell2.font = header_font
                header_cell2.fill = header_fill
                header_cell2.alignment = center_alignment
                
                # Información del móvil
                movil_data = [
                    ('Patente', movil_info.get('patente', 'N/A')),
                    ('Alias', movil_info.get('alias', 'N/A')),
                    ('Marca', movil_info.get('marca', 'N/A')),
                    ('Modelo', movil_info.get('modelo', 'N/A')),
                    ('', ''),  # Línea en blanco
                ]
                
                # Estadísticas del recorrido
                stats_data = [
                    ('Duración (minutos)', estadisticas.get('duracion_minutos', 0)),
                    ('Distancia (km)', estadisticas.get('distancia_km', 0)),
                    ('Velocidad Máxima (km/h)', estadisticas.get('velocidad_maxima', 0)),
                    ('Velocidad Promedio (km/h)', estadisticas.get('velocidad_promedio', 0)),
                    ('Puntos GPS', estadisticas.get('puntos_gps', 0)),
                    ('Detenciones', estadisticas.get('detenciones', 0)),
                    ('Tiempo Detenido (minutos)', estadisticas.get('tiempo_detenido_minutos', 0)),
                    ('Tiempo en Movimiento (minutos)', estadisticas.get('tiempo_movimiento_minutos', 0)),
                ]
                
                # Combinar datos
                all_data = movil_data + stats_data
                
                for row, (stat, value) in enumerate(all_data, 2):
                    ws_stats.cell(row=row, column=1, value=stat)
                    ws_stats.cell(row=row, column=2, value=value)
                
                # Ajustar ancho de columnas
                ws_stats.column_dimensions['A'].width = 25
                ws_stats.column_dimensions['B'].width = 20
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Crear respuesta HTTP
            from django.http import HttpResponse
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="recorrido_gps_{len(posiciones)}_posiciones.xlsx"'
            
            return response
            
        except Exception as e:
            return Response({'error': f'Error generando Excel: {str(e)}'}, status=500)
    
    @action(detail=False, methods=['post'])
    def geocodificar_posicion(self, request):
        """Geocodificar una posición individual"""
        lat = request.data.get('lat')
        lon = request.data.get('lon')
        posicion_id = request.data.get('posicion_id')
        
        if not lat or not lon:
            return Response({'error': 'Latitud y longitud son requeridas'}, status=400)
        
        try:
            from gps.services import GeocodingService
            
            geocoding_service = GeocodingService()
            resultado = geocoding_service.geocodificar_coordenadas(float(lat), float(lon))
            
            # Extraer la dirección del resultado
            if resultado and 'direccion_formateada' in resultado:
                direccion = resultado['direccion_formateada']
            elif resultado and 'display_name' in resultado:
                direccion = resultado['display_name']
            else:
                direccion = f"Coordenadas: {lat}, {lon}"
            
            # Si se proporciona posicion_id, actualizar la base de datos
            if posicion_id:
                try:
                    posicion = Posicion.objects.get(id=posicion_id)
                    posicion.direccion = direccion
                    posicion.save()
                except Posicion.DoesNotExist:
                    pass  # Continuar sin error si no se encuentra la posición
            
            return Response({
                'lat': lat,
                'lon': lon,
                'direccion': direccion
            })
            
        except Exception as e:
            return Response({'error': f'Error geocodificando: {str(e)}'}, status=500)
    
    @action(detail=False, methods=['post'])
    def geocodificar_recorrido(self, request):
        """Geocodificar múltiples posiciones"""
        posiciones = request.data.get('posiciones', [])
        
        if not posiciones:
            return Response({'error': 'No hay posiciones para geocodificar'}, status=400)
        
        try:
            from gps.services import GeocodingService
            import time
            
            geocoding_service = GeocodingService()
            direcciones = []
            geocodificadas = 0
            
            for posicion_data in posiciones:
                lat = posicion_data.get('lat')
                lon = posicion_data.get('lon')
                posicion_id = posicion_data.get('id')
                
                if lat and lon:
                    try:
                        resultado = geocoding_service.geocodificar_coordenadas(float(lat), float(lon))
                        
                        # Extraer la dirección del resultado
                        if resultado and 'direccion_formateada' in resultado:
                            direccion = resultado['direccion_formateada']
                        elif resultado and 'display_name' in resultado:
                            direccion = resultado['display_name']
                        else:
                            direccion = f"Coordenadas: {lat}, {lon}"
                        
                        # Actualizar en la base de datos si se proporciona ID
                        if posicion_id:
                            try:
                                posicion = Posicion.objects.get(id=posicion_id)
                                posicion.direccion = direccion
                                posicion.save()
                                print(f"✅ Dirección guardada para Posicion ID {posicion_id}: {direccion[:50]}...")
                            except Posicion.DoesNotExist:
                                print(f"⚠️ Posicion con ID {posicion_id} no encontrada en la base de datos")
                                pass
                        
                        direcciones.append(direccion)
                        geocodificadas += 1
                        
                        # Pequeña pausa para no sobrecargar el servicio
                        time.sleep(0.1)
                        
                    except Exception as e:
                        direcciones.append(f'Error: {str(e)}')
                else:
                    direcciones.append('Coordenadas inválidas')
            
            return Response({
                'geocodificadas': geocodificadas,
                'total': len(posiciones),
                'direcciones': direcciones
            })
            
        except Exception as e:
            return Response({'error': f'Error en geocodificación masiva: {str(e)}'}, status=500)


def recorridos_frontend(request):
    """Renderiza la página de recorridos"""
    return render(request, 'recorridos/index.html')


# API endpoints para control del receptor
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny


@api_view(['GET'])
@permission_classes([AllowAny])
def diagnostico_performance(request):
    """
    Endpoint de diagnóstico para identificar cuellos de botella
    """
    import time
    from django.db import connection
    from django.core.cache import cache
    
    resultados = {
        'timestamp': time.time(),
        'tests': {}
    }
    
    # Test 1: Conexión a BD
    start = time.time()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        resultados['tests']['db_connection'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'ok'
        }
    except Exception as e:
        resultados['tests']['db_connection'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'error',
            'error': str(e)
        }
    
    # Test 2: Query simple de móviles
    start = time.time()
    try:
        count = Movil.objects.filter(activo=True).count()
        resultados['tests']['query_moviles'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'ok',
            'count': count
        }
    except Exception as e:
        resultados['tests']['query_moviles'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'error',
            'error': str(e)
        }
    
    # Test 3: Cache
    start = time.time()
    try:
        cache.set('test_diagnostico', 'ok', 10)
        valor = cache.get('test_diagnostico')
        resultados['tests']['cache'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'ok' if valor == 'ok' else 'error'
        }
    except Exception as e:
        resultados['tests']['cache'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'error',
            'error': str(e)
        }
    
    # Test 4: Query de posiciones (limitada)
    start = time.time()
    try:
        count = Posicion.objects.filter(is_valid=True)[:100].count()
        resultados['tests']['query_posiciones'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'ok',
            'count': count
        }
    except Exception as e:
        resultados['tests']['query_posiciones'] = {
            'tiempo': round(time.time() - start, 3),
            'status': 'error',
            'error': str(e)
        }
    
    resultados['tiempo_total'] = round(sum(t['tiempo'] for t in resultados['tests'].values()), 3)
    
    return Response(resultados)


def comunicaciones_frontend(request):
    """Renderiza la página de comunicaciones"""
    return render(request, 'comunicaciones/index.html')


def configuraciones_receptor_frontend(request):
    """Renderiza la página de configuraciones de receptor"""
    return render(request, 'comunicaciones/configuraciones.html')


def estadisticas_receptor_frontend(request):
    """Renderiza la página de estadísticas de receptor"""
    return render(request, 'comunicaciones/estadisticas.html')


def logs_receptor_frontend(request):
    """Renderiza la página de logs de receptores"""
    return render(request, 'comunicaciones/logs.html')


def mapa_frontend(request):
    """Renderiza la página del mapa con la flota en tiempo real"""
    return render(request, 'mapa/index.html')


def reportes_frontend(request):
    """Renderiza la página de reportes"""
    return render(request, 'reportes/index.html')


def controles_neo_frontend(request):
    """Renderiza la página de Controles NEO"""
    return render(request, 'controles-neo/index.html')


# API endpoints para control del receptor
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny

@api_view(['GET'])
@permission_classes([AllowAny])
def receiver_stats(request):
    """Obtener estadísticas del receptor TCP"""
    from gps.receiver_manager import get_receiver_stats
    
    stats = get_receiver_stats()
    if stats:
        return Response(stats)
    else:
        return Response({
            'running': False,
            'message': 'No hay receptores activos'
        })


@api_view(['POST'])
@permission_classes([AllowAny])
def receiver_start(request):
    """Iniciar un receptor en un puerto específico"""
    from gps.receiver_manager import start_receiver
    
    host = request.data.get('host', '0.0.0.0')
    port = request.data.get('port', 5003)
    
    try:
        port = int(port)
    except (ValueError, TypeError):
        return Response({'error': 'Puerto inválido'}, status=400)
    
    result = start_receiver(host=host, port=port)
    
    if result['success']:
        return Response(result)
    else:
        return Response(result, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
def receiver_stop(request):
    """Detener un receptor en un puerto específico"""
    from gps.receiver_manager import stop_receiver
    
    port = request.data.get('port')
    
    if port is None:
        return Response({'error': 'Puerto es requerido'}, status=400)
    
    try:
        port = int(port)
    except (ValueError, TypeError):
        return Response({'error': 'Puerto inválido'}, status=400)
    
    result = stop_receiver(port=port)
    
    if result['success']:
        return Response(result)
    else:
        return Response(result, status=400)


@api_view(['GET'])
@permission_classes([AllowAny])
def test_api(request):
    """Endpoint de prueba para verificar que la API funciona"""
    return Response({'status': 'ok', 'message': 'API funcionando correctamente'})


@api_view(['GET'])
@permission_classes([AllowAny])
def receiver_logs(request):
    """Obtener lista de archivos de log de receptores"""
    from gps.logging_manager import logging_manager
    
    port = request.GET.get('port')
    log_files = logging_manager.get_log_files(port=int(port) if port else None)
    print("📂 receiver_logs - base log dir:", logging_manager.base_log_dir)
    print("📂 receiver_logs - archivos encontrados:", [str(p) for p in log_files])
    
    # Convertir Path objects a diccionarios
    files_data = []
    for log_file in log_files:
        try:
            stat = log_file.stat()
            files_data.append({
                'path': str(log_file),
                'name': log_file.name,
                'size': stat.st_size,
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'port': log_file.parent.name.split('_')[1] if '_' in log_file.parent.name else 'Unknown'
            })
        except Exception as e:
            continue
    
    return Response(files_data)


@api_view(['GET'])
@permission_classes([AllowAny])
def receiver_log_content(request):
    """Obtener contenido de un archivo de log específico"""
    from gps.logging_manager import logging_manager
    
    log_path = request.GET.get('path')
    lines = request.GET.get('lines', 100)
    
    if not log_path:
        return Response({'error': 'Path es requerido'}, status=400)
    
    try:
        lines = int(lines)
    except (ValueError, TypeError):
        lines = 100
    
    content = logging_manager.get_log_content(log_path, lines)
    
    # Si el contenido es un mensaje de error, devolverlo como error
    if content.startswith('Error') or content.startswith('Archivo de log no encontrado'):
        return Response({'error': content}, status=404)
    
    return Response({'content': content})